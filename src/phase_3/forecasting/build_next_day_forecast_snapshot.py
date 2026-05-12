from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from email.utils import parsedate_to_datetime
from pathlib import Path

import joblib
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import requests
import urllib3
from catboost import CatBoostRegressor
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]

DATA_DIR = PROJECT_ROOT / "data" / "phase_3" / "forecasting"
EXTERNAL_DIR = DATA_DIR / "external"
PLOTS_DIR = PROJECT_ROOT / "pictures" / "phase_3" / "forecasting"
ALL_FIGURES_DIR = PROJECT_ROOT / "pictures" / "phase_3" / "all_figures"

for directory in [DATA_DIR, EXTERNAL_DIR, PLOTS_DIR, ALL_FIGURES_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

KOSTT_PAGE_URL = "https://kostt.com/Transparency/BasicMarketDataOnGeneration"
KOSTT_EXCEL_URL = (
    "https://kostt.com/Content/ViewFiles/Transparency/BasicMarketDataOnGeneration/"
    "Plani%20i%20prodhimit%20total%20te%20energjise%20elektrike%20per%20diten%20ne%20vijim.xlsx"
)

WEATHER_API_URL = "https://api.open-meteo.com/v1/forecast"
PRISHTINA_LAT = 42.6629
PRISHTINA_LON = 21.1655

TARGET = "pm25"

SCALER_PATH = PROJECT_ROOT / "models" / "scaler.pkl"
SELECTED_DATA_PATH = PROJECT_ROOT / "data" / "phase_1" / "4E_selected_dataset.csv"
RAW_HISTORY_PATH = PROJECT_ROOT / "data" / "phase_1" / "2D_validated_final_dataset.csv"
CAP_REFERENCE_PATH = PROJECT_ROOT / "data" / "phase_1" / "4A_outliers_handled.csv"

PHASE3_MODEL_PATH = PROJECT_ROOT / "models" / "phase_3" / "catboost_tuned" / "catboost_phase3_tuned_model.cbm"
PHASE3_FEATURES_PATH = PROJECT_ROOT / "models" / "phase_3" / "catboost_tuned" / "catboost_phase3_feature_columns.pkl"
PHASE2_MODEL_PATH = PROJECT_ROOT / "models" / "catboost_model" / "catboost_pm25_model.cbm"
PHASE2_FEATURES_PATH = PROJECT_ROOT / "models" / "catboost_model" / "catboost_feature_columns.pkl"

LOG1P_FEATURES = {TARGET, "pollution_stagnation_index", "rain"}


def save_phase3_figure(filename: str, **kwargs: object) -> None:
    plt.savefig(PLOTS_DIR / filename, **kwargs)
    plt.savefig(ALL_FIGURES_DIR / filename, **kwargs)


def request_headers() -> dict[str, str]:
    return {"User-Agent": "Mozilla/5.0 (phase-3-air-pollution-project)"}


def download_kostt_excel() -> tuple[Path, dict[str, str]]:
    output_path = EXTERNAL_DIR / "kostt_generation_plan_next_day_snapshot.xlsx"
    try:
        response = requests.get(KOSTT_EXCEL_URL, headers=request_headers(), timeout=40)
    except requests.exceptions.SSLError:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        response = requests.get(KOSTT_EXCEL_URL, headers=request_headers(), timeout=40, verify=False)
    response.raise_for_status()
    output_path.write_bytes(response.content)
    return output_path, dict(response.headers)


def parse_last_modified(headers: dict[str, str]) -> date | None:
    value = headers.get("Last-Modified") or headers.get("last-modified")
    if not value:
        return None
    try:
        return parsedate_to_datetime(value).date()
    except Exception:
        return None


def swap_month_day(value: date) -> date | None:
    try:
        return date(value.year, value.day, value.month)
    except ValueError:
        return None


def parse_kostt_excel(path: Path, headers: dict[str, str]) -> dict[str, object]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    dates: list[date] = []
    numbers: list[float] = []

    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, datetime):
                dates.append(value.date())
            elif isinstance(value, date):
                dates.append(value)
            elif isinstance(value, (int, float)) and float(value) > 100:
                numbers.append(float(value))

    if not dates:
        raise ValueError("KOSTT forecast date was not found in the Excel snapshot.")
    if not numbers:
        raise ValueError("KOSTT forecast MWh value was not found in the Excel snapshot.")

    raw_forecast_date = dates[-1]
    forecast_mwh = numbers[-1]
    last_modified_date = parse_last_modified(headers)

    corrected_date = raw_forecast_date
    correction_note = "no correction"
    if last_modified_date is not None:
        expected_next_day = last_modified_date + timedelta(days=1)
        swapped = swap_month_day(raw_forecast_date)
        if raw_forecast_date != expected_next_day and swapped == expected_next_day:
            corrected_date = expected_next_day
            correction_note = "corrected by swapping day/month using HTTP Last-Modified + 1 day"
        elif raw_forecast_date != expected_next_day:
            correction_note = "date differs from HTTP Last-Modified + 1 day; kept Excel date"

    snapshot = {
        "source_page": KOSTT_PAGE_URL,
        "source_file": KOSTT_EXCEL_URL,
        "downloaded_at": datetime.now().isoformat(timespec="seconds"),
        "http_last_modified": headers.get("Last-Modified") or headers.get("last-modified", ""),
        "raw_forecast_date": raw_forecast_date.isoformat(),
        "forecast_date": corrected_date.isoformat(),
        "date_correction_note": correction_note,
        "generation_forecast_mwh": forecast_mwh,
        "local_excel_path": str(path),
    }
    pd.DataFrame([snapshot]).to_csv(DATA_DIR / "kostt_next_day_generation_snapshot.csv", index=False)
    return snapshot


def fetch_weather_forecast(forecast_date: date) -> pd.DataFrame:
    params = {
        "latitude": PRISHTINA_LAT,
        "longitude": PRISHTINA_LON,
        "hourly": "temperature_2m,rain,relative_humidity_2m,wind_speed_10m,wind_direction_10m",
        "timezone": "Europe/Belgrade",
        "start_date": forecast_date.isoformat(),
        "end_date": forecast_date.isoformat(),
    }
    response = requests.get(WEATHER_API_URL, params=params, headers=request_headers(), timeout=40)
    response.raise_for_status()
    payload = response.json()

    with open(EXTERNAL_DIR / "open_meteo_next_day_weather_snapshot.json", "w", encoding="utf-8") as file:
        json.dump(payload, file, indent=2)

    hourly = payload["hourly"]
    weather = pd.DataFrame({
        "timestamp": pd.to_datetime(hourly["time"], errors="coerce"),
        "temperature_2m": hourly["temperature_2m"],
        "rain": hourly["rain"],
        "relative_humidity_2m": hourly["relative_humidity_2m"],
        "wind_speed_10m": hourly["wind_speed_10m"],
        "wind_direction_10m": hourly["wind_direction_10m"],
    })
    weather = weather.dropna(subset=["timestamp"]).reset_index(drop=True)
    weather.to_csv(EXTERNAL_DIR / "open_meteo_next_day_weather_snapshot.csv", index=False)
    if len(weather) != 24:
        raise ValueError(f"Expected 24 hourly weather rows, got {len(weather)}.")
    return weather


def find_col(columns: list[str], prefix: str) -> str:
    for column in columns:
        if column.startswith(prefix):
            return column
    raise KeyError(f"Column starting with '{prefix}' was not found.")


def load_model_bundle() -> tuple[CatBoostRegressor, list[str], str]:
    if PHASE3_MODEL_PATH.exists() and PHASE3_FEATURES_PATH.exists():
        model_path = PHASE3_MODEL_PATH
        features_path = PHASE3_FEATURES_PATH
        source = "phase_3_tuned_catboost"
    else:
        model_path = PHASE2_MODEL_PATH
        features_path = PHASE2_FEATURES_PATH
        source = "phase_2_catboost_fallback"

    model = CatBoostRegressor()
    model.load_model(str(model_path))
    feature_cols = joblib.load(features_path)
    return model, feature_cols, source


def build_hourly_generation_profile(forecast_date: date, total_mwh: float) -> pd.DataFrame:
    history = pd.read_csv(RAW_HISTORY_PATH)
    history["datetime"] = pd.to_datetime(history["datetime"], errors="coerce")
    history = history.dropna(subset=["datetime"])
    history["month"] = history["datetime"].dt.month
    history["hour"] = history["datetime"].dt.hour

    month_history = history[history["month"] == forecast_date.month]
    if month_history.empty:
        month_history = history

    hourly_mean = month_history.groupby("hour")["total_generation_mw"].mean().reindex(range(24))
    hourly_mean = hourly_mean.fillna(hourly_mean.mean())
    weights = hourly_mean / hourly_mean.sum()

    timestamps = pd.date_range(start=pd.Timestamp(forecast_date), periods=24, freq="h")
    generation = pd.DataFrame({
        "timestamp": timestamps,
        "hour": range(24),
        "generation_profile_weight": weights.values,
    })
    generation["total_generation_mw"] = generation["generation_profile_weight"] * float(total_mwh)
    generation.to_csv(DATA_DIR / "kostt_hourly_generation_profile_from_daily_total.csv", index=False)
    return generation


def load_scaled_history() -> pd.DataFrame:
    df = pd.read_csv(SELECTED_DATA_PATH)
    time_col = find_col(df.columns.tolist(), "datetime")
    df[time_col] = pd.to_datetime(df[time_col], errors="coerce")
    df = df.dropna(subset=[time_col]).sort_values(time_col).drop_duplicates(time_col).reset_index(drop=True)
    df["timestamp"] = df[time_col]
    df["month"] = df["timestamp"].dt.month
    df["hour"] = df["timestamp"].dt.hour
    return df


def build_pm25_seed_profile(history_scaled: pd.DataFrame) -> tuple[pd.Series, float]:
    profile = history_scaled.groupby(["month", "hour"])[TARGET].median()
    global_median = float(history_scaled[TARGET].median())
    return profile, global_median


def typical_pm25_model_space(ts: pd.Timestamp, profile: pd.Series, global_median: float) -> float:
    key = (int(ts.month), int(ts.hour))
    if key in profile.index:
        return float(profile.loc[key])
    return global_median


def load_cap_ranges() -> dict[str, tuple[float, float]]:
    if not CAP_REFERENCE_PATH.exists():
        return {}
    df = pd.read_csv(CAP_REFERENCE_PATH)
    ranges: dict[str, tuple[float, float]] = {}
    for column in df.select_dtypes(include=[np.number]).columns:
        ranges[column] = (float(df[column].min()), float(df[column].max()))
    return ranges


def is_log1p_feature(feature_name: str) -> bool:
    return feature_name == TARGET or feature_name == "pollution_stagnation_index" or feature_name.startswith("rain")


def preprocess_value(feature_name: str, value: float, cap_ranges: dict[str, tuple[float, float]]) -> float:
    value = float(value)
    if feature_name in cap_ranges:
        lower, upper = cap_ranges[feature_name]
        value = float(np.clip(value, lower, upper))
    if is_log1p_feature(feature_name):
        value = float(np.log1p(max(value, 0.0)))
    return value


def scale_feature(feature_name: str, raw_value: float, scaler: object, cap_ranges: dict[str, tuple[float, float]]) -> float:
    feature_names = list(getattr(scaler, "feature_names_in_", []))
    if feature_name not in feature_names:
        return float(raw_value)
    idx = feature_names.index(feature_name)
    prepared = preprocess_value(feature_name, raw_value, cap_ranges)
    return float((prepared - scaler.mean_[idx]) / scaler.scale_[idx])


def inverse_scale_pm25(value: float, scaler: object) -> float:
    feature_names = list(getattr(scaler, "feature_names_in_", []))
    if TARGET not in feature_names:
        return float(value)
    idx = feature_names.index(TARGET)
    log_value = float(value) * scaler.scale_[idx] + scaler.mean_[idx]
    return float(np.expm1(log_value))


def build_feature_row(
    ts: pd.Timestamp,
    controls: dict[str, float],
    lag1_model: float,
    lag24_model: float,
    feature_cols: list[str],
    scaler: object,
    cap_ranges: dict[str, tuple[float, float]],
) -> pd.DataFrame:
    generation = float(controls["total_generation_mw"])
    temp = float(controls["temperature_2m"])
    rain = max(0.0, float(controls["rain"]))
    humidity = float(np.clip(controls["relative_humidity_2m"], 0.0, 100.0))
    wind_speed = max(0.0, float(controls["wind_speed_10m"]))
    wind_dir = float(controls["wind_direction_10m"]) % 360.0

    hour_sin = float(np.sin(2 * np.pi * ts.hour / 24.0))
    hour_cos = float(np.cos(2 * np.pi * ts.hour / 24.0))
    month_sin = float(np.sin(2 * np.pi * ts.month / 12.0))
    month_cos = float(np.cos(2 * np.pi * ts.month / 12.0))
    wind_rad = np.deg2rad(wind_dir)
    wind_x = float(wind_speed * np.cos(wind_rad))
    wind_y = float(wind_speed * np.sin(wind_rad))
    stagnation = float(generation / (wind_speed + 0.1))

    raw_values = {}
    for feature in feature_cols:
        if feature == "hour_sin":
            raw_values[feature] = hour_sin
        elif feature == "hour_cos":
            raw_values[feature] = hour_cos
        elif feature == "month_sin":
            raw_values[feature] = month_sin
        elif feature == "month_cos":
            raw_values[feature] = month_cos
        elif feature == "pollution_stagnation_index":
            raw_values[feature] = stagnation
        elif feature == "wind_x_vector":
            raw_values[feature] = wind_x
        elif feature == "wind_y_vector":
            raw_values[feature] = wind_y
        elif feature == "total_generation_mw":
            raw_values[feature] = generation
        elif feature.startswith("temperature_2m"):
            raw_values[feature] = temp
        elif feature.startswith("rain"):
            raw_values[feature] = rain
        elif feature.startswith("relative_humidity_2m"):
            raw_values[feature] = humidity
        elif feature.startswith("wind_direction_10m"):
            raw_values[feature] = wind_dir
        elif feature.startswith("wind_speed_10m"):
            raw_values[feature] = wind_speed

    row = {}
    for feature in feature_cols:
        if feature == "pm25_lag_1":
            row[feature] = lag1_model
        elif feature == "pm25_lag_24":
            row[feature] = lag24_model
        else:
            row[feature] = scale_feature(feature, raw_values.get(feature, 0.0), scaler, cap_ranges)
    return pd.DataFrame([row], columns=feature_cols)


def risk_category(pm25_mean: float) -> str:
    if pm25_mean < 10:
        return "Low"
    if pm25_mean < 20:
        return "Moderate"
    if pm25_mean < 35:
        return "Elevated"
    return "High"


def build_forecast_snapshot(
    kostt_snapshot: dict[str, object],
    weather: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    model, feature_cols, model_source = load_model_bundle()
    scaler = joblib.load(SCALER_PATH)
    history_scaled = load_scaled_history()
    pm25_profile, global_pm25 = build_pm25_seed_profile(history_scaled)
    cap_ranges = load_cap_ranges()

    forecast_date = date.fromisoformat(str(kostt_snapshot["forecast_date"]))
    total_mwh = float(kostt_snapshot["generation_forecast_mwh"])
    generation = build_hourly_generation_profile(forecast_date, total_mwh)

    frame = weather.merge(generation[["timestamp", "total_generation_mw", "generation_profile_weight"]], on="timestamp", how="inner")
    if len(frame) != 24:
        raise ValueError(f"Weather/generation merge produced {len(frame)} rows instead of 24.")

    history_model = {
        pd.Timestamp(row["timestamp"]): float(row[TARGET])
        for _, row in history_scaled[["timestamp", TARGET]].iterrows()
    }

    rows = []
    for _, item in frame.iterrows():
        ts = pd.Timestamp(item["timestamp"])
        lag1_ts = ts - pd.Timedelta(hours=1)
        lag24_ts = ts - pd.Timedelta(hours=24)
        lag1_model = history_model.get(lag1_ts, typical_pm25_model_space(lag1_ts, pm25_profile, global_pm25))
        lag24_model = history_model.get(lag24_ts, typical_pm25_model_space(lag24_ts, pm25_profile, global_pm25))

        controls = {
            "total_generation_mw": float(item["total_generation_mw"]),
            "temperature_2m": float(item["temperature_2m"]),
            "rain": float(item["rain"]),
            "relative_humidity_2m": float(item["relative_humidity_2m"]),
            "wind_speed_10m": float(item["wind_speed_10m"]),
            "wind_direction_10m": float(item["wind_direction_10m"]),
        }
        row = build_feature_row(ts, controls, lag1_model, lag24_model, feature_cols, scaler, cap_ranges)
        pred_model = float(model.predict(row)[0])
        pred_real = inverse_scale_pm25(pred_model, scaler)
        history_model[ts] = pred_model

        rows.append({
            "timestamp": ts,
            "forecast_date": forecast_date.isoformat(),
            "pm25_forecast": pred_real,
            "pm25_model_space": pred_model,
            "total_generation_mw": controls["total_generation_mw"],
            "temperature_2m": controls["temperature_2m"],
            "rain": controls["rain"],
            "relative_humidity_2m": controls["relative_humidity_2m"],
            "wind_speed_10m": controls["wind_speed_10m"],
            "wind_direction_10m": controls["wind_direction_10m"],
            "lag_seed_method": "recursive plus historical month-hour median",
            "model_source": model_source,
        })

    hourly_df = pd.DataFrame(rows)
    hourly_df.to_csv(DATA_DIR / "next_day_pm25_hourly_forecast_snapshot.csv", index=False)

    daily_summary = pd.DataFrame([{
        "forecast_date": forecast_date.isoformat(),
        "model_source": model_source,
        "generation_forecast_mwh": total_mwh,
        "pm25_mean_forecast": float(hourly_df["pm25_forecast"].mean()),
        "pm25_max_forecast": float(hourly_df["pm25_forecast"].max()),
        "pm25_min_forecast": float(hourly_df["pm25_forecast"].min()),
        "peak_hour": str(hourly_df.loc[hourly_df["pm25_forecast"].idxmax(), "timestamp"]),
        "risk_category": risk_category(float(hourly_df["pm25_forecast"].mean())),
        "weather_temperature_mean": float(hourly_df["temperature_2m"].mean()),
        "weather_rain_sum": float(hourly_df["rain"].sum()),
        "weather_wind_speed_mean": float(hourly_df["wind_speed_10m"].mean()),
        "snapshot_note": "Stored forecast snapshot; online refresh is optional for demo day.",
    }])
    daily_summary.to_csv(DATA_DIR / "next_day_pm25_daily_summary_snapshot.csv", index=False)

    return hourly_df, daily_summary


def save_snapshot_plot(hourly_df: pd.DataFrame, daily_summary: pd.DataFrame) -> None:
    forecast_date = daily_summary["forecast_date"].iloc[0]
    risk = daily_summary["risk_category"].iloc[0]
    mean_pm25 = daily_summary["pm25_mean_forecast"].iloc[0]

    fig, ax1 = plt.subplots(figsize=(13, 5))
    ax1.plot(hourly_df["timestamp"], hourly_df["pm25_forecast"], marker="o", color="#4C78A8", label="PM2.5 forecast")
    ax1.axhline(mean_pm25, color="#D62728", linestyle="--", linewidth=1.5, label="Daily mean")
    ax1.set_ylabel("PM2.5")
    ax1.set_title(f"Stored next-day PM2.5 forecast snapshot for {forecast_date} ({risk})")
    ax1.tick_params(axis="x", rotation=25)

    ax2 = ax1.twinx()
    ax2.bar(hourly_df["timestamp"], hourly_df["total_generation_mw"], width=0.03, color="#72B7B2", alpha=0.25, label="Distributed generation")
    ax2.set_ylabel("Generation MW")

    handles1, labels1 = ax1.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(handles1 + handles2, labels1 + labels2, loc="upper left")

    plt.tight_layout()
    save_phase3_figure("next_day_pm25_forecast_snapshot.png", dpi=300, bbox_inches="tight")
    plt.close()


def main() -> None:
    print("=" * 88)
    print("PHASE 3 :: STORED NEXT-DAY FORECAST SNAPSHOT")
    print("=" * 88)
    excel_path, headers = download_kostt_excel()
    kostt_snapshot = parse_kostt_excel(excel_path, headers)
    forecast_date = date.fromisoformat(str(kostt_snapshot["forecast_date"]))
    weather = fetch_weather_forecast(forecast_date)
    hourly_df, daily_summary = build_forecast_snapshot(kostt_snapshot, weather)
    save_snapshot_plot(hourly_df, daily_summary)

    run_info = {
        "kostt_snapshot": kostt_snapshot,
        "outputs": {
            "daily_summary": str(DATA_DIR / "next_day_pm25_daily_summary_snapshot.csv"),
            "hourly_forecast": str(DATA_DIR / "next_day_pm25_hourly_forecast_snapshot.csv"),
            "generation_snapshot": str(DATA_DIR / "kostt_next_day_generation_snapshot.csv"),
            "weather_snapshot": str(EXTERNAL_DIR / "open_meteo_next_day_weather_snapshot.csv"),
            "plot": str(PLOTS_DIR / "next_day_pm25_forecast_snapshot.png"),
            "all_figures_plot": str(ALL_FIGURES_DIR / "next_day_pm25_forecast_snapshot.png"),
        },
        "all_figures_dir": str(ALL_FIGURES_DIR),
        "offline_demo_policy": "Use the saved snapshot files first; refresh from KOSTT/Open-Meteo only when desired.",
    }
    with open(DATA_DIR / "next_day_forecast_snapshot_run_info.json", "w", encoding="utf-8") as file:
        json.dump(run_info, file, indent=2, default=str)

    print(daily_summary.to_string(index=False))
    print("\nSaved next-day forecast snapshot to:", DATA_DIR)


if __name__ == "__main__":
    main()
